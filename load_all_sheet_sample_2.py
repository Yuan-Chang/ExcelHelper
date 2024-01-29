import utils.Utils as utils
from utils.modules import Sheet
import pandas as pd
from openpyxl import load_workbook
import utils.openpyxl_helper as helper

sample_file = "samples/loadAllSheetsSample/sample3.xlsx"
macro_sample_file = "samples/loadAllSheetsSample/empty_macro.xlsm"
output_folder = "samples/loadAllSheetsSample/output/"

sample_copy_file = output_folder + utils.get_file_name_from_file_path(sample_file)
macro_sample_copy_file = output_folder + utils.get_file_name_from_file_path(macro_sample_file)
# sample_copy_file = utils.replace_extension(sample_copy_file, "xlsx")

result_file = output_folder + "result.xlsm"

# utils.quit_excel()

utils.delete_directory(output_folder)
utils.create_directory(output_folder)

# Create a sample copy
utils.create_file_copy(sample_file, sample_copy_file)
utils.create_file_copy(macro_sample_file, macro_sample_copy_file)

# Prepare the sheet
# utils.create_values_only_excel_file(input_file=sample_copy_file, output_file=result_file)

df1 = pd.DataFrame({'Data': ['a', 'b', 'c', 'd']})
df2 = pd.DataFrame({'Data': [1, 2, 3, 4]})
df3 = pd.DataFrame({'Data': [1.1, 1.2, 1.3, 1.4]})

list = [Sheet("week1", df1), Sheet("week2", df2), Sheet("week3", df3)]

wb = load_workbook(sample_copy_file)
macro_wb = load_workbook(macro_sample_file, keep_vba=True)

helper.merge_two_workbook(macro_wb, wb)

helper.write_sheet_to_worksheet(macro_wb, list)

helper.copy_worksheet(macro_wb["Sheet1_or"], macro_wb["week1"], copy_value=False)

macro_wb.save(result_file)

