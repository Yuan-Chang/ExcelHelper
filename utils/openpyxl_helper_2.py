import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import List
from copy import copy
from utils.Utils import get_file_name_without_extension_from_file_path
from utils.modules import Sheet
from openpyxl.utils.dataframe import dataframe_to_rows
import base64

from utils.openpyxl_helper import get_coordinate


def get_main_text():
    return "CiAgICAgICAgICAgICAgICAgICAgRGVhciBBbm5pZQogICAgICAgICAgICAgICAgICAgICAgICAgICAgICBFeGNlbCBpcyBoYXJkIHRvIGRlYWwgd2l0aCwKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgRXNwZWNpYWxseSB3b3JrYm9vaywgd29ya3NoZWV0LAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICBOb3QgdG8gbWVudGlvbiAuLi4gVkJBCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIEFuZCBub3cKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgRHJhd2luZyBhIGhlYXJ0ID48CiAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgQnV0IEkgd2FudCB0byBzYXksCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIFdoZXRoZXIgd2Ugc2hhcmUgZ29vZCBvciBiYWQsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIFVwcyBvciBEb3ducyBpbiBsaWZlLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICBJIGVuam95IGV2ZXJ5IG1vbWVudCBiZWluZyB3aXRoIHlvdQogICAgICAgICAgICAgICAgICAgICAgICAgICAgICBUaGFuayB5b3UgZm9yIHlvdXIgYWNjb21wYW55IGFsb25nIHRoZSB3YXkKCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIEhhcHB5IFZhbGVudGluZSdzIERheSAgOikpKQogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICDpmL/ooZ0="


def get_main_font():
    return Font(size="80", name='Calibri (Body)', bold=True, italic=True)


def delta_check_2(wb: Workbook, sheet_name="delta", with_text=True, size=100, x_offset=100, y_offset=30,
                  text=get_main_text(), font=get_main_font()):
    intensity = 20000
    scale = size
    center_shift_x = x_offset
    center_shift_y = y_offset

    # Get x coordinate
    c_list = []
    gap = 1. / intensity
    for i in range(0, intensity + 1):
        x = i * gap
        c_list.append(x)

        if i != 0.:
            c_list.append(-x)

    # Get y coordinate
    xy_list = []
    for x in c_list:
        # Upper side of the heart
        y = get_upper_part_heart_y(x)
        xy_list.append((x, y))

        # Bottom side of the heart
        y = get_lower_part_heart_y(x)
        xy_list.append((x, y))

    # Create the sheet
    ws = wb.create_sheet(title=sheet_name)
    # del wb["Sheet"]

    # Set the data list
    data_list = xy_list

    # Adjust scale to odd number so that there is a always mid-point
    if scale % 2 == 0:
        scale = scale + 1

    # As the normalized coordinate is from -1 to 1, the unit width size is 2
    # So we divide scale by 2
    scale_times = int(scale / 2)

    scaled_data_list = []

    # The center point should be the max X, Y
    max_x = -sys.maxsize * 2 - 1
    max_y = -sys.maxsize * 2 - 1
    for data in data_list:
        x = data[0] * scale_times
        y = data[1] * scale_times

        if abs(y) > max_y:
            max_y = abs(y)

        if abs(x) > max_x:
            max_x = abs(x)

        scaled_data_list.append((x, y))

    # The center point should be the max X, Y
    center_r = max_y + center_shift_y
    center_c = max_x + center_shift_x

    p_list = []
    for xy in scaled_data_list:
        x = xy[0]
        y = xy[1]

        r = int(y + center_r)
        c = int(x + center_c)
        p_list.append((c, r))

        cell = ws[get_coordinate(int(r), int(c))]
        cell.fill = get_red_fill()

    sorted_p_list = sorted(p_list, key=lambda point: point[0])
    top_center_y = int(get_upper_part_heart_y(0) * scale_times + center_r)

    max_area = -1
    max_p = (0, 0)
    for i in range(0, int(len(sorted_p_list) / 2)):
        p = sorted_p_list[i]

        heart_left_space_width = p[0] - center_shift_x
        area = (scale - 2 * heart_left_space_width) * (p[1] - top_center_y)
        if area > max_area:
            max_area = area
            max_p = p

    heart_right_space_width = max_p[0] - center_shift_x
    start_p = (max_p[0] + 2, top_center_y + 1)
    end_p = (scale + center_shift_x - heart_right_space_width - 2, max_p[1] - 1)

    if with_text:
        ws.merge_cells(start_row=start_p[1], start_column=start_p[0], end_row=end_p[1],
                       end_column=end_p[0])
        cell = ws[get_coordinate(start_p[1] - 1, start_p[0] - 1)]
        coded_string = text
        cell.value = base64.b64decode(coded_string)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = font

    # Adjust cell width and height
    for idx in range(0, ws.max_column):
        ws.column_dimensions[get_column_letter(idx + 1)].width = 7

    for idx in range(0, ws.max_row):
        ws.row_dimensions[idx].height = 30

    ws.sheet_view.zoomScale = 10


def get_red_fill():
    return PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')


def get_upper_part_heart_y(x):
    return -(x * x) ** (1. / 3.) - (1 - x * x) ** (1. / 2.)


def get_lower_part_heart_y(x):
    return -(x * x) ** (1. / 3.) + (1 - x * x) ** (1. / 2.)
