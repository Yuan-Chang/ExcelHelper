import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
import shutil
from typing import List
from utils.modules import Sheet
import pandas as pd
from copy import copy
from utils.Utils import get_file_name_without_extension_from_file_path
from utils.modules import Sheet
from openpyxl.utils.dataframe import dataframe_to_rows


class DataColumn:
    def __init__(self, position, key, data_values: List):
        self.position = position
        self.key = key
        self.data_values = data_values


def copy_worksheet(source_ws, target_ws, copy_value: bool = True):
    for row in source_ws.iter_rows():
        for cell in row:
            if copy_value:
                target_ws[cell.coordinate].value = cell.value

            target_ws[cell.coordinate].font = copy(cell.font)
            target_ws[cell.coordinate].border = copy(cell.border)
            target_ws[cell.coordinate].fill = copy(cell.fill)
            target_ws[cell.coordinate].number_format = cell.number_format
            target_ws[cell.coordinate].protection = copy(cell.protection)
            target_ws[cell.coordinate].alignment = copy(cell.alignment)
            target_ws[cell.coordinate].comment = cell.comment

    # Copy cell width and height
    for idx, rd in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx] = copy(rd)

    for idx, rd in source_ws.column_dimensions.items():
        target_ws.column_dimensions[idx] = copy(rd)

    for merged_cell in source_ws.merged_cells:
        target_ws.merge_cells(f"{merged_cell}")


def merge_multiple_excels_to_one_excel(input_files: List, output_file):
    output_wb = Workbook()

    for file_name in input_files:
        source_wb = load_workbook(file_name)
        number_of_sheets = len(source_wb.sheetnames)

        for sheet in source_wb:
            file_name = get_file_name_without_extension_from_file_path(file_name)
            sheet_name = f"{file_name}_{sheet.title}"

            # if only 1 sheet, use file name instead
            if number_of_sheets == 1:
                sheet_name = file_name

            output_ws = output_wb.create_sheet(sheet_name)
            copy_worksheet(sheet, output_ws)

    # Delete the default sheet
    del output_wb["Sheet"]
    output_wb.save(output_file)


# merge wb2 into wb1
def merge_two_workbook(wb1: Workbook, wb2: Workbook):
    source_wb = wb2

    for sheet in source_wb:
        sheet_name = sheet.title

        output_ws = wb1.create_sheet(sheet_name)
        copy_worksheet(sheet, output_ws)


def write_sheet_to_worksheet(wb: Workbook, sheets: List[Sheet]):
    for sheet in sheets:
        ws = wb.create_sheet(sheet.name)
        rows = dataframe_to_rows(sheet.data_frame, index=False)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)


def save_file_as_xlsm(workbook: Workbook, result_file, empty_macro_file):
    macro_wb = load_workbook(empty_macro_file, keep_vba=True)
    merge_two_workbook(macro_wb, workbook)
    macro_wb.save(result_file)


def delta_check(current_ws: Worksheet, previous_ws: Worksheet, delta_ws: Worksheet, start_point: List,
                header_count: int = 2):
    start_row_index = start_point[0] - 1
    start_column_index = column_index_from_string(start_point[1]) - 1

    current_header_keys = get_header_keys(start_row_index, start_column_index, current_ws, header_count=header_count)
    current_data_columns = get_data_columns(start_row_index, start_column_index, current_ws, header_count=header_count)

    previous_header_keys = get_header_keys(start_row_index, start_column_index, previous_ws, header_count=header_count)
    previous_data_columns = get_data_columns(start_row_index, start_column_index, previous_ws,
                                             header_count=header_count)

    # h1 h2 "" h3
    # h2 "" h3
    header_dict = {}
    for header_column_index, header in enumerate(previous_header_keys):
        if header != "":
            header_dict[header] = header_column_index

    # Calculate and write the delta cell value
    for header_column_index, header in enumerate(current_header_keys):

        if header in header_dict:
            # value check
            current_data_column = current_data_columns[header_column_index]
            previous_data_column = previous_data_columns[header_dict[header]]

            for data_row_index, current_cell in enumerate(current_data_column):

                previous_cell = previous_data_column[data_row_index]

                result_value = f"='{current_ws.title}'!{current_cell.coordinate}-'{previous_ws.title}'!{previous_cell.coordinate}"

                delta_ws[current_cell.coordinate].value = result_value


    # Copy the rest
    for r_index, row in enumerate(current_ws.iter_rows()):
        for c_index, cell in enumerate(row):
            if r_index < start_row_index + header_count or c_index < start_column_index:
                delta_ws[cell.coordinate].value = cell.value

    copy_worksheet(current_ws, delta_ws, copy_value=False)


def get_header_keys(start_row_index: int, start_column_index: int, ws: Worksheet, header_count: int = 2):
    header_row_range = range(start_row_index, start_row_index + header_count)
    header_column_range = range(start_column_index, ws.max_column)

    key_dict = {}
    key_list = []

    # for row_index in header_row_range:
    for column_index in header_column_range:

        key = ""
        for row_index in header_row_range:
            cell = ws[get_coordinate(row_index, column_index)]
            merged_cell_value = get_merged_cell_value(ws, cell)

            if merged_cell_value is None:
                merged_cell_value = ""

            if key == "":
                key = merged_cell_value
            else:
                key = f"{key}_{merged_cell_value}"

        # Handle duplicate key
        # if duplicate exist, we append the count to the end
        # ex. key, key_1, key_2
        if key != "":
            if key in key_dict:
                count = key_dict[key]
                key = f"{key}_{count}"
                key_dict[key] = count + 1
            else:
                key_dict[key] = 1

        key_list.append(key)
    return key_list


def get_data_columns(start_row_index: int, start_column_index: int, ws: Worksheet, header_count: int = 2):
    header_row_range = range(start_row_index + header_count, ws.max_row)
    header_column_range = range(start_column_index, ws.max_column)

    column_list = []

    # for row_index in header_row_range:
    for column_index in header_column_range:

        column = []
        for row_index in header_row_range:
            cell = ws[get_coordinate(row_index, column_index)]
            column.append(cell)

        column_list.append(column)

    return column_list


def get_coordinate(r_index, c_index):
    col = get_column_letter(c_index + 1)
    return f"{col}{r_index + 1}"


# if the cell is part of a merged cell, return the first cell value
# If not, return cell value
def get_merged_cell_value(ws: Worksheet, cell):
    for merged_cell in ws.merged_cells.ranges:
        if cell.coordinate in merged_cell:
            pair = merged_cell.left[0]
            return ws.cell(row=pair[0], column=pair[1]).value
    return cell.value

# def testMerge(row, column):
#     cell = sheet.cell(row, column)
#     for mergedCell in sheet.merged_cells.ranges:
#         if (cell.coordinate in mergedCell):
#             return True
#     return False
