from openpyxl import load_workbook
import utils.Utils as utils
import utils.openpyxl_helper as openpyxl_helper

sample_file = "samples/versionCheck/sample.xlsm"
output_folder = "samples/versionCheck/output/"

sample_copy_file = output_folder + utils.get_file_name_from_file_path(sample_file)

result_file = output_folder + "result.xlsm"

utils.delete_directory(output_folder)
utils.create_directory(output_folder)

# Create a sample copy
utils.create_file_copy(sample_file, sample_copy_file)

list = [["Sheet1", "Sheet2", [5, "C"]]]

wb = load_workbook(sample_copy_file, keep_vba=True)

for data in list:
    ws1 = wb[data[0]]
    ws2 = wb[data[1]]
    delta_ws = wb.create_sheet(f"{data[0]} delta")
    start_point = data[2]
    openpyxl_helper.delta_check(ws1, ws2, delta_ws, start_point)

wb.save(result_file)
